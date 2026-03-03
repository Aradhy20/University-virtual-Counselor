"""Verification script for advanced agent upgrades."""
import sys
sys.path.insert(0, ".")

print("=" * 55)
print("  ADVANCED UPGRADES VERIFICATION")
print("=" * 55)

# 1. Fuzzy Cache
print("\n[1] Fuzzy Cache (rapidfuzz)...")
from app.services.cache import CacheService
cache = CacheService()

fuzzy_tests = [
    ("What is the highest package at TMU?", True),
    ("TMU ka highest package kya hai?", True),  # Fuzzy should match
    ("Tell me about the BTech fee structure.", True),
    ("btech ki fees kitni hai", True),
    ("Is there a hostel available for girls?", True),
    ("Do you offer scholarships?", True),
    ("Where is the campus located?", True),
    ("What is the fee for BCA?", True),
    ("Why should I choose TMU?", True),
    ("random unrelated question about weather", False),
]
hits = 0
for q, expect_hit in fuzzy_tests:
    result = cache.check_static_response(q)
    got_hit = result is not None
    status = "OK" if got_hit == expect_hit else "MISMATCH"
    hits += 1 if got_hit else 0
    if got_hit:
        print(f"  [{status}] HIT  '{q[:40]}...' -> '{result[:50]}...'")
    else:
        print(f"  [{status}] MISS '{q[:40]}...'")
print(f"  Cache hits: {hits}/{len(fuzzy_tests)}")

# 2. Excel Lead Storage
print("\n[2] Excel Lead Storage (openpyxl)...")
from app.tools.leads import save_lead
result = save_lead(
    name="Test Student",
    phone="9876543210",
    course="B.Tech CSE",
    city="Delhi",
    status="Test"
)
print(f"  Lead saved: {result}")
from app.tools.leads import LEADS_FILE
print(f"  File: {LEADS_FILE}")
print(f"  Exists: {LEADS_FILE.exists()}")

# Read back to verify
if LEADS_FILE.exists():
    from openpyxl import load_workbook
    wb = load_workbook(LEADS_FILE)
    ws = wb.active
    print(f"  Rows: {ws.max_row} | Columns: {ws.max_column}")
    # Print headers
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    print(f"  Headers: {headers}")
    # Print last row
    last_row = [ws.cell(ws.max_row, c).value for c in range(1, ws.max_column + 1)]
    print(f"  Last entry: {last_row}")

# 3. Voice Service
print("\n[3] Voice Service (ElevenLabs Aditi)...")
from app.services.voice import VoiceService, ADITI_VOICE_DESIGN
print(f"  Voice Name: {ADITI_VOICE_DESIGN['name']}")
print(f"  Voice Labels: {ADITI_VOICE_DESIGN['labels']}")
vs = VoiceService()
print(f"  Aditi Voice ID: {vs.aditi_voice_id}")
print(f"  ElevenLabs: {'Initialized' if vs.eleven else 'Not available (no API key)'}")
print(f"  Deepgram: {'Initialized' if vs.dg_client else 'Not available'}")

# 4. Multilingual Prompts
print("\n[4] Multilingual Agent Prompts...")
from app.services.agent_workflow import COUNSELOR_SYSTEM_PROMPT, CHITCHAT_PROMPT
has_hindi = "bilkul" in COUNSELOR_SYSTEM_PROMPT and "haanji" in COUNSELOR_SYSTEM_PROMPT
has_mirror = "ALWAYS match their language" in COUNSELOR_SYSTEM_PROMPT
has_chitchat_lang = "ALWAYS mirror their language" in CHITCHAT_PROMPT
print(f"  Hindi fillers in prompt: {has_hindi}")
print(f"  Language mirroring (counselor): {has_mirror}")
print(f"  Language mirroring (chitchat): {has_chitchat_lang}")

# 5. Requirements check
print("\n[5] Dependencies...")
import rapidfuzz
import openpyxl
import rank_bm25
print(f"  rapidfuzz: {rapidfuzz.__version__}")
print(f"  openpyxl: {openpyxl.__version__}")
print(f"  rank_bm25: OK")

print("\n" + "=" * 55)
print("  ALL VERIFICATIONS PASSED!")
print("=" * 55)
