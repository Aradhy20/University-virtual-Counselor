"""
Lead Storage — Excel (.xlsx) format using openpyxl.

Columns: Date | Name | Mobile Number | Interested Course
File: data/leads.xlsx
"""
import logging
from pathlib import Path
from datetime import datetime

logger = logging.getLogger("riya.leads")

LEADS_FILE = Path(__file__).parent.parent.parent / "data" / "leads.xlsx"


def save_lead(name: str, phone: str = "Unknown", course: str = "Unknown",
              city: str = "Unknown", status: str = "New") -> bool:
    """
    Appends a new lead to data/leads.xlsx.
    Creates the file with formatted headers if it doesn't exist.
    """
    try:
        from openpyxl import Workbook, load_workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

        # Ensure directory exists
        if not LEADS_FILE.parent.exists():
            LEADS_FILE.parent.mkdir(parents=True, exist_ok=True)

        if LEADS_FILE.exists():
            wb = load_workbook(LEADS_FILE)
            ws = wb.active
        else:
            # Create new workbook with formatted header
            wb = Workbook()
            ws = wb.active
            ws.title = "TMU Leads"

            # Header row
            headers = ["Date", "Name", "Mobile Number", "Interested Course", "City", "Status"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF", size=12)
                cell.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = Border(
                    bottom=Side(style="thin"),
                    right=Side(style="thin")
                )

            # Set column widths
            ws.column_dimensions["A"].width = 20  # Date
            ws.column_dimensions["B"].width = 25  # Name
            ws.column_dimensions["C"].width = 18  # Mobile
            ws.column_dimensions["D"].width = 25  # Course
            ws.column_dimensions["E"].width = 18  # City
            ws.column_dimensions["F"].width = 12  # Status

        # Append new lead
        row_data = [
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            name or "Unknown",
            phone or "Unknown",
            course or "Unknown",
            city or "Unknown",
            status
        ]
        ws.append(row_data)

        # Style the new row
        new_row = ws.max_row
        for col in range(1, 7):
            cell = ws.cell(row=new_row, column=col)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(
                bottom=Side(style="thin", color="D3D3D3"),
                right=Side(style="thin", color="D3D3D3")
            )

        wb.save(LEADS_FILE)
        logger.info(f"Lead saved to Excel: {name} | {course} | {phone}")
        return True

    except Exception as e:
        logger.error(f"Failed to save lead to Excel: {e}")
        # Fallback to CSV if openpyxl fails
        return _fallback_csv(name, phone, course, city, status)


def _fallback_csv(name, phone, course, city, status):
    """CSV fallback in case openpyxl has issues."""
    import csv
    csv_path = LEADS_FILE.parent / "leads_backup.csv"
    try:
        file_exists = csv_path.exists()
        with open(csv_path, "a", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            if not file_exists:
                writer.writerow(["Date", "Name", "Mobile Number", "Interested Course", "City", "Status"])
            writer.writerow([
                datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                name or "Unknown", phone or "Unknown",
                course or "Unknown", city or "Unknown", status
            ])
        logger.info(f"Lead saved to CSV fallback: {name}")
        return True
    except Exception as e:
        logger.error(f"CSV fallback also failed: {e}")
        return False
